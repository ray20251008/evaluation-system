import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def build_blank_assessment_excel():
    wb = openpyxl.Workbook()
    default_sheet = wb.active
    
    # Colors
    c_primary = "1E3A8A"      # Navy Blue
    c_primary_light = "DBEAFE"
    c_accent = "0D9488"       # Teal
    c_accent_light = "CCFBF1"
    c_subtle = "F8FAFC"       # Slate 50
    c_border = "CBD5E1"       # Slate 300
    c_green_light = "DCFCE7"

    thin_border = Border(
        left=Side(style='thin', color=c_border),
        right=Side(style='thin', color=c_border),
        top=Side(style='thin', color=c_border),
        bottom=Side(style='thin', color=c_border)
    )
    thick_bottom = Border(
        left=Side(style='thin', color=c_border),
        right=Side(style='thin', color=c_border),
        top=Side(style='thin', color=c_border),
        bottom=Side(style='medium', color=c_primary)
    )

    # -------------------------------------------------------------
    # 1. Sheet 1: 使用說明與評分標準
    # -------------------------------------------------------------
    ws_guide = wb.create_sheet(title="評分標準與說明")
    ws_guide.views.sheetView[0].showGridLines = True
    
    ws_guide["A1"] = "116年度身心障礙者服務衛教執行計畫暨成效評量手冊 - 評分標準與使用說明 (空白範本)"
    ws_guide["A1"].font = Font(name="微軟正黑體", size=16, bold=True, color="FFFFFF")
    ws_guide["A1"].fill = PatternFill(start_color=c_primary, end_color=c_primary, fill_type="solid")
    ws_guide["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_guide.merge_cells("A1:F2")

    guide_headers = ["分數", "評分等級", "判定標準與具體表現說明", "學習階段判定", "代表色彩", "備註說明"]
    for col_idx, h in enumerate(guide_headers, 1):
        cell = ws_guide.cell(row=4, column=col_idx, value=h)
        cell.font = Font(name="微軟正黑體", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=c_accent, end_color=c_accent, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    standards = [
        (5, "可獨立完成", "不需要任何提示，能主動且正確無誤地完成完整步驟。", "🟢 獨立自主階段 (4.5~5.0分)", "綠色", "達到完全自立水準"),
        (4, "口語提示", "支持者給予口頭提醒（如：「夾的動作呢？」）即可正確完成。", "🔵 口語提示階段 (3.5~4.4分)", "藍色", "達到基本自立達標標準"),
        (3, "肢體協助", "需要支持者給予肢體引導或輕碰示範（如牽手引導角度）才能完成。", "🟡 肢體協助階段 (2.5~3.4分)", "黃色", "需日常持續操作練習"),
        (2, "部分協助", "需支持者協助操作超過一半或關鍵步驟才能完成。", "🟡 部分協助階段 (2.0~2.4分)", "橙色", "需拆解步驟個別引導"),
        (1, "大量協助", "大部分步驟由支持者代為操作，使用者僅能參與極少部分。", "🟠 大量協助階段 (1.0~1.9分)", "紅橙色", "需支持者全程密集陪伴"),
        (0, "無法完成", "即使提供大量協助仍無法完成或拒絕配合操作。", "🔴 引導探索階段 (0.0~0.9分)", "紅色", "需建立動機與基礎習慣")
    ]

    for r_idx, row_data in enumerate(standards, 5):
        for c_idx, val in enumerate(row_data, 1):
            cell = ws_guide.cell(row=r_idx, column=c_idx, value=val)
            cell.font = Font(name="微軟正黑體", size=10)
            cell.alignment = Alignment(horizontal="center" if c_idx in [1, 2, 4, 5] else "left", vertical="center", wrap_text=True)
            cell.border = thin_border
            if r_idx % 2 == 0:
                cell.fill = PatternFill(start_color=c_subtle, end_color=c_subtle, fill_type="solid")

    ws_guide.cell(row=12, column=1, value="◆ 填寫方式指引").font = Font(name="微軟正黑體", size=12, bold=True, color=c_primary)
    instructions = [
        ("步驟 1", "請直接切換至各季評量打分表（Q1、Q2、Q3、Q4）。"),
        ("步驟 2", "在「學員姓名」欄位輸入您的學員名字（可自行複製新增多位學員）。"),
        ("步驟 3", "在指標 1 ～ 10 欄位輸入 0 ～ 5 分的評量得分。"),
        ("步驟 4", "系統將自動計算【總得分】、【平均分】、【學習階段】、【達成率】、【進步分數】與【進步率】！")
    ]
    for idx, (step_lbl, desc) in enumerate(instructions, 13):
        ws_guide.cell(row=idx, column=1, value=step_lbl).alignment = Alignment(horizontal="center")
        ws_guide.cell(row=idx, column=2, value=desc)
        ws_guide.merge_cells(start_row=idx, start_column=2, end_row=idx, end_column=6)
        for col in range(1, 7):
            c = ws_guide.cell(row=idx, column=col)
            c.font = Font(name="微軟正黑體", size=10)
            c.border = thin_border

    ws_guide.column_dimensions['A'].width = 14
    ws_guide.column_dimensions['B'].width = 18
    ws_guide.column_dimensions['C'].width = 38
    ws_guide.column_dimensions['D'].width = 25
    ws_guide.column_dimensions['E'].width = 12
    ws_guide.column_dimensions['F'].width = 25

    # -------------------------------------------------------------
    # 2. Quarterly Sheets Data (Q1, Q2, Q3, Q4) - BLANK SCORES
    # -------------------------------------------------------------
    quarters_data = [
        {
            "id": "Q1",
            "title": "第一季 (1~3月)：健康衛生與日常照護",
            "items": [
                "1. 能說出需要洗手的時機（如飯前、便後、擤鼻涕後等）",
                "2. 能正確執行內外夾弓大立腕洗手步驟",
                "3. 能將雙手搓洗至乾淨並以流動水沖洗乾淨",
                "4. 能正確拿取牙刷及牙膏並擠出適當量",
                "5. 能維持適當角度（貝氏刷牙法）刷牙",
                "6. 能依序清潔牙齒各區域（內側、外側、咬合面）",
                "7. 能用流動水清潔傷口及周圍髒污",
                "8. 能正確撕開並使用OK繃／透氣膠帶包紮傷口",
                "9. 止鼻血時能保持頭部微前傾並正確加壓止血",
                "10. 止血或傷口護理時，能知道何時需尋求支持者協助"
            ]
        },
        {
            "id": "Q2",
            "title": "第二季 (4~6月)：健康飲食與身體監測",
            "items": [
                "1. 能分辨紅（少吃）、黃（適量）、綠（天天吃）燈食物",
                "2. 能從菜單或實物中挑選澱粉類（全穀雜糧）食物",
                "3. 能挑選蛋白質（蛋豆魚肉）食物",
                "4. 能主動選擇或夾取蔬菜類食物",
                "5. 能依據均衡飲食原則組合一份完整餐點",
                "6. 能正確操作額溫槍並對準額頭量測",
                "7. 能正確配戴壓脈帶並操作電子血壓計",
                "8. 能脫鞋站穩量測體重並正確讀取數值",
                "9. 能正確夾好血氧機並保持靜止量測",
                "10. 能看懂自己的健康數據並說出異常時需注意的地方"
            ]
        },
        {
            "id": "Q3",
            "title": "第三季 (7~9月)：緊急事故與健康安全",
            "items": [
                "1. 能辨識活動區域周遭的危險物品（剪刀、熱水等）",
                "2. 遇到緊急或意外狀況時能立即停止原本手邊活動",
                "3. 遇到他人跌倒或發作能協助移開周遭危險物品",
                "4. 能知道同儕癲癇發作時「不可強行壓制或塞東西入嘴」",
                "5. 能聽從指示維持現場安全與通風環境",
                "6. 發生意外事件時能知道需要立刻尋求成人/支持者協助",
                "7. 遇到危急事件時能大聲呼救或通知他人",
                "8. 能說出或指出緊急電話「119」",
                "9. 能清楚說出自己的姓名、所在位置等基本報案資訊",
                "10. 遇到自己無法處理的情況能主動向外界求助"
            ]
        },
        {
            "id": "Q4",
            "title": "第四季 (10~12月)：身體自主與健康自我倡議",
            "items": [
                "1. 能辨識並察覺自己的身體出現不舒服或異常",
                "2. 能準確指出或觸摸自己不舒服的身體部位",
                "3. 能用言語或圖卡表達疼痛（如刺痛、悶痛）或不適感受",
                "4. 能知道身體不適時需第一時間告知支持者或家屬",
                "5. 能說出或指出需要尋求的醫療協助（如看醫生、吃藥）",
                "6. 能正確選擇適合的日用／夜用型衛生用品",
                "7. 能正確完成衛生用品的黏貼與更換完整流程",
                "8. 能妥善包裝並丟棄使用過的衛生用品至垃圾桶",
                "9. 就醫看診時能回答醫師的基本問診問題",
                "10. 能主動向醫師或支持者提出自己的健康疑問"
            ]
        }
    ]

    # Pre-configure 8 blank learner rows
    blank_learners = [f"學員 {i+1}" for i in range(8)]

    for q_info in quarters_data:
        ws_q = wb.create_sheet(title=f"{q_info['id']}評量打分")
        ws_q.views.sheetView[0].showGridLines = True

        # Title
        ws_q["A1"] = f"{q_info['title']} - 學員前測／課堂平時／後測評量打分表"
        ws_q["A1"].font = Font(name="微軟正黑體", size=14, bold=True, color="FFFFFF")
        ws_q["A1"].fill = PatternFill(start_color=c_primary, end_color=c_primary, fill_type="solid")
        ws_q["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws_q.merge_cells("A1:S2")

        # Instructions banner
        ws_q["A3"] = "【評分提示等級】5:可獨立完成 ｜ 4:口語提示 ｜ 3:肢體協助 ｜ 2:部分協助 ｜ 1:大量協助 ｜ 0:無法完成 (在指標 1~10 欄位填入 0~5 分)"
        ws_q["A3"].font = Font(name="微軟正黑體", size=10, bold=True, color="1E3A8A")
        ws_q["A3"].fill = PatternFill(start_color=c_primary_light, end_color=c_primary_light, fill_type="solid")
        ws_q["A3"].alignment = Alignment(horizontal="left", vertical="center")
        ws_q.merge_cells("A3:S3")

        # Table Header
        ws_q["A4"] = "學員姓名"
        ws_q["A4"].font = Font(name="微軟正黑體", size=11, bold=True, color="FFFFFF")
        ws_q["A4"].fill = PatternFill(start_color=c_accent, end_color=c_accent, fill_type="solid")
        ws_q["A4"].alignment = Alignment(horizontal="center", vertical="center")
        ws_q.merge_cells("A4:A5")

        ws_q["B4"] = "評量階段"
        ws_q["B4"].font = Font(name="微軟正黑體", size=11, bold=True, color="FFFFFF")
        ws_q["B4"].fill = PatternFill(start_color=c_accent, end_color=c_accent, fill_type="solid")
        ws_q["B4"].alignment = Alignment(horizontal="center", vertical="center")
        ws_q.merge_cells("B4:B5")

        for i in range(1, 11):
            col_letter = get_column_letter(2 + i)
            cell = ws_q[f"{col_letter}4"]
            cell.value = f"指標 {i}"
            cell.font = Font(name="微軟正黑體", size=10, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=c_accent, end_color=c_accent, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

            sub_cell = ws_q[f"{col_letter}5"]
            sub_cell.value = f"(0~5分)"
            sub_cell.font = Font(name="微軟正黑體", size=8, color="FFFFFF")
            sub_cell.fill = PatternFill(start_color=c_accent, end_color=c_accent, fill_type="solid")
            sub_cell.alignment = Alignment(horizontal="center", vertical="center")

        calc_headers = [
            ("M", "總得分", "(滿分50)"),
            ("N", "平均分", "(0~5分)"),
            ("O", "學習階段判定", "(等級標記)"),
            ("P", "達成率", "(%)"),
            ("Q", "進步分數", "(後-前)"),
            ("R", "進步率", "(%)"),
            ("S", "成長百分點", "(%pt)")
        ]
        for col_let, h_main, h_sub in calc_headers:
            c1 = ws_q[f"{col_let}4"]
            c1.value = h_main
            c1.font = Font(name="微軟正黑體", size=10, bold=True, color="FFFFFF")
            c1.fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
            c1.alignment = Alignment(horizontal="center", vertical="center")

            c2 = ws_q[f"{col_let}5"]
            c2.value = h_sub
            c2.font = Font(name="微軟正黑體", size=8, color="FFFFFF")
            c2.fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
            c2.alignment = Alignment(horizontal="center", vertical="center")

        start_row = 6
        for l_idx, learner_name in enumerate(blank_learners):
            r_pre = start_row + l_idx * 3
            r_mid = r_pre + 1
            r_post = r_pre + 2

            # Name cell
            ws_q[f"A{r_pre}"] = learner_name
            ws_q[f"A{r_pre}"].font = Font(name="微軟正黑體", size=11, bold=True)
            ws_q[f"A{r_pre}"].alignment = Alignment(horizontal="center", vertical="center")
            ws_q.merge_cells(f"A{r_pre}:A{r_post}")

            # Stage labels
            ws_q[f"B{r_pre}"] = "1. 前測 (第1月)"
            ws_q[f"B{r_mid}"] = "2. 課堂 (第2月)"
            ws_q[f"B{r_post}"] = "3. 後測 (第3月)"
            for r, stage_fill in [(r_pre, "F8FAFC"), (r_mid, "F1F5F9"), (r_post, "E2E8F0")]:
                c = ws_q[f"B{r}"]
                c.font = Font(name="微軟正黑體", size=9, bold=(r == r_post))
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.fill = PatternFill(start_color=stage_fill, end_color=stage_fill, fill_type="solid")

            # Blank score cells with clean borders
            for i in range(10):
                col_let = get_column_letter(3 + i)
                for r in [r_pre, r_mid, r_post]:
                    cell = ws_q[f"{col_let}{r}"]
                    cell.value = ""  # Clean blank!
                    cell.font = Font(name="微軟正黑體", size=10)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border

            # Error-safe Formulas for rows (Display blank if no scores entered yet)
            for r in [r_pre, r_mid, r_post]:
                ws_q[f"M{r}"] = f'=IF(COUNT(C{r}:L{r})>0, SUM(C{r}:L{r}), "")'
                ws_q[f"N{r}"] = f'=IF(COUNT(C{r}:L{r})>0, AVERAGE(C{r}:L{r}), "")'
                ws_q[f"O{r}"] = f'=IF(N{r}="","",IF(N{r}>=4.5,"🟢 獨立完成",IF(N{r}>=3.5,"🔵 口語提示",IF(N{r}>=2.0,"🟡 肢體/部分協助","🟠 大量協助"))))'
                ws_q[f"P{r}"] = f'=IF(ISNUMBER(M{r}), M{r}/50, "")'

                ws_q[f"M{r}"].font = Font(name="微軟正黑體", size=10, bold=True)
                ws_q[f"N{r}"].font = Font(name="微軟正黑體", size=10)
                ws_q[f"O{r}"].font = Font(name="微軟正黑體", size=9, bold=True)
                ws_q[f"P{r}"].font = Font(name="微軟正黑體", size=10)
                ws_q[f"N{r}"].number_format = '0.0'
                ws_q[f"P{r}"].number_format = '0.0%'

                for col in ["M", "N", "O", "P"]:
                    ws_q[f"{col}{r}"].alignment = Alignment(horizontal="center", vertical="center")
                    ws_q[f"{col}{r}"].border = thin_border

            # Improvement formulas (Display only when post-test has score)
            ws_q[f"Q{r_pre}"] = f'=IF(AND(ISNUMBER(M{r_pre}), ISNUMBER(M{r_post})), M{r_post}-M{r_pre}, "")'
            ws_q[f"R{r_pre}"] = f'=IF(AND(ISNUMBER(M{r_pre}), ISNUMBER(M{r_post})), IF(M{r_pre}>0, (M{r_post}-M{r_pre})/M{r_pre}, (M{r_post}-M{r_pre})/50), "")'
            ws_q[f"S{r_pre}"] = f'=IF(AND(ISNUMBER(P{r_pre}), ISNUMBER(P{r_post})), P{r_post}-P{r_pre}, "")'

            ws_q.merge_cells(f"Q{r_pre}:Q{r_post}")
            ws_q.merge_cells(f"R{r_pre}:R{r_post}")
            ws_q.merge_cells(f"S{r_pre}:S{r_post}")

            for col in ["Q", "R", "S"]:
                cell = ws_q[f"{col}{r_pre}"]
                cell.font = Font(name="微軟正黑體", size=10, bold=True, color="166534")
                cell.fill = PatternFill(start_color=c_green_light, end_color=c_green_light, fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

            ws_q[f"Q{r_pre}"].number_format = '+0;-0;0'
            ws_q[f"R{r_pre}"].number_format = '+0.0%;-0.0%;0.0%'
            ws_q[f"S{r_pre}"].number_format = '+0.0%;-0.0%;0.0%'

            for col_idx in range(1, 20):
                cell_name = get_column_letter(col_idx)
                for r in [r_pre, r_mid, r_post]:
                    ws_q[f"{cell_name}{r}"].border = thin_border

        # Class Averages Row
        r_avg = start_row + len(blank_learners) * 3
        ws_q[f"A{r_avg}"] = "全班平均常模"
        ws_q[f"A{r_avg}"].font = Font(name="微軟正黑體", size=10, bold=True, color="FFFFFF")
        ws_q[f"A{r_avg}"].fill = PatternFill(start_color=c_primary, end_color=c_primary, fill_type="solid")
        ws_q[f"A{r_avg}"].alignment = Alignment(horizontal="center", vertical="center")
        ws_q.merge_cells(f"A{r_avg}:B{r_avg}")

        for i in range(10):
            col_let = get_column_letter(3 + i)
            cell = ws_q[f"{col_let}{r_avg}"]
            cell.value = f'=IFERROR(AVERAGE({col_let}6:{col_let}{r_avg-1}), "-")'
            cell.font = Font(name="微軟正黑體", size=10, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(start_color=c_primary_light, end_color=c_primary_light, fill_type="solid")
            cell.border = thick_bottom

        ws_q[f"M{r_avg}"] = f'=IFERROR(AVERAGE(M6:M{r_avg-1}), "-")'
        ws_q[f"N{r_avg}"] = f'=IFERROR(AVERAGE(N6:N{r_avg-1}), "-")'
        ws_q[f"O{r_avg}"] = "-"
        ws_q[f"P{r_avg}"] = f'=IFERROR(AVERAGE(P6:P{r_avg-1}), "-")'
        ws_q[f"Q{r_avg}"] = f'=IFERROR(AVERAGE(Q6:Q{r_avg-1}), "-")'
        ws_q[f"R{r_avg}"] = f'=IFERROR(AVERAGE(R6:R{r_avg-1}), "-")'
        ws_q[f"S{r_avg}"] = f'=IFERROR(AVERAGE(S6:S{r_avg-1}), "-")'

        for col in ["M", "N", "O", "P", "Q", "R", "S"]:
            cell = ws_q[f"{col}{r_avg}"]
            cell.font = Font(name="微軟正黑體", size=10, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(start_color=c_primary_light, end_color=c_primary_light, fill_type="solid")
            cell.border = thick_bottom

        # Indicators Reference below
        r_ref = r_avg + 3
        ws_q.cell(row=r_ref, column=1, value=f"【{q_info['id']} 檢核指標項目清單參照】").font = Font(name="微軟正黑體", size=11, bold=True, color=c_primary)
        for i, item_text in enumerate(q_info['items'], 1):
            r_item = r_ref + i
            ws_q.cell(row=r_item, column=1, value=f"指標 {i}").alignment = Alignment(horizontal="center")
            ws_q.cell(row=r_item, column=2, value=item_text)
            ws_q.merge_cells(start_row=r_item, start_column=2, end_row=r_item, end_column=12)
            for col in range(1, 13):
                ws_q.cell(row=r_item, column=col).font = Font(name="微軟正黑體", size=9)
                ws_q.cell(row=r_item, column=col).border = thin_border

        # Adjust widths
        ws_q.column_dimensions['A'].width = 14
        ws_q.column_dimensions['B'].width = 16
        for i in range(1, 11):
            ws_q.column_dimensions[get_column_letter(2 + i)].width = 9
        ws_q.column_dimensions['M'].width = 10
        ws_q.column_dimensions['N'].width = 9
        ws_q.column_dimensions['O'].width = 18
        ws_q.column_dimensions['P'].width = 10
        ws_q.column_dimensions['Q'].width = 11
        ws_q.column_dimensions['R'].width = 11
        ws_q.column_dimensions['S'].width = 13

    # -------------------------------------------------------------
    # 3. Sheet: 年度全期統計總表 (空白)
    # -------------------------------------------------------------
    ws_summary = wb.create_sheet(title="年度成果統計總表")
    ws_summary.views.sheetView[0].showGridLines = True

    ws_summary["A1"] = "116年度全年度四季前後測成績與進步成效總表 (空白範本)"
    ws_summary["A1"].font = Font(name="微軟正黑體", size=14, bold=True, color="FFFFFF")
    ws_summary["A1"].fill = PatternFill(start_color=c_primary, end_color=c_primary, fill_type="solid")
    ws_summary["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_summary.merge_cells("A1:P2")

    sum_headers = [
        ("A", "學員姓名"),
        ("B", "Q1前測"), ("C", "Q1課堂"), ("D", "Q1後測"),
        ("E", "Q2前測"), ("F", "Q2課堂"), ("G", "Q2後測"),
        ("H", "Q3前測"), ("I", "Q3課堂"), ("J", "Q3後測"),
        ("K", "Q4前測"), ("L", "Q4課堂"), ("M", "Q4後測"),
        ("N", "年度年初基準"), ("O", "年度年末總結"), ("P", "年度總進步率")
    ]
    for col_let, h in sum_headers:
        c = ws_summary[f"{col_let}4"]
        c.value = h
        c.font = Font(name="微軟正黑體", size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill(start_color=c_accent, end_color=c_accent, fill_type="solid")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border

    for idx, l_name in enumerate(blank_learners, 5):
        ws_summary[f"A{idx}"] = l_name
        ws_summary[f"B{idx}"] = f"='Q1評量打分'!M{6 + (idx-5)*3}"
        ws_summary[f"C{idx}"] = f"='Q1評量打分'!M{7 + (idx-5)*3}"
        ws_summary[f"D{idx}"] = f"='Q1評量打分'!M{8 + (idx-5)*3}"

        ws_summary[f"E{idx}"] = f"='Q2評量打分'!M{6 + (idx-5)*3}"
        ws_summary[f"F{idx}"] = f"='Q2評量打分'!M{7 + (idx-5)*3}"
        ws_summary[f"G{idx}"] = f"='Q2評量打分'!M{8 + (idx-5)*3}"

        ws_summary[f"H{idx}"] = f"='Q3評量打分'!M{6 + (idx-5)*3}"
        ws_summary[f"I{idx}"] = f"='Q3評量打分'!M{7 + (idx-5)*3}"
        ws_summary[f"J{idx}"] = f"='Q3評量打分'!M{8 + (idx-5)*3}"

        ws_summary[f"K{idx}"] = f"='Q4評量打分'!M{6 + (idx-5)*3}"
        ws_summary[f"L{idx}"] = f"='Q4評量打分'!M{7 + (idx-5)*3}"
        ws_summary[f"M{idx}"] = f"='Q4評量打分'!M{8 + (idx-5)*3}"

        ws_summary[f"N{idx}"] = f"=B{idx}"
        ws_summary[f"O{idx}"] = f"=M{idx}"
        ws_summary[f"P{idx}"] = f'=IF(AND(ISNUMBER(N{idx}), ISNUMBER(O{idx}), N{idx}>0), (O{idx}-N{idx})/N{idx}, "")'

        for col_idx in range(1, 17):
            col_let = get_column_letter(col_idx)
            c = ws_summary[f"{col_let}{idx}"]
            c.font = Font(name="微軟正黑體", size=10)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = thin_border
            if col_idx in [14, 15]:
                c.fill = PatternFill(start_color=c_primary_light, end_color=c_primary_light, fill_type="solid")
            elif col_idx == 16:
                c.fill = PatternFill(start_color=c_green_light, end_color=c_green_light, fill_type="solid")
                c.font = Font(name="微軟正黑體", size=10, bold=True, color="166534")
                c.number_format = '+0.0%;-0.0%;0.0%'

    for col_idx in range(1, 17):
        col_let = get_column_letter(col_idx)
        ws_summary.column_dimensions[col_let].width = 12
    ws_summary.column_dimensions['A'].width = 14
    ws_summary.column_dimensions['P'].width = 15

    # Remove default sheet
    if default_sheet in wb.worksheets:
        wb.remove(default_sheet)

    out_file = "116年度衛教與學習成效評量分析表_空白範本檔.xlsx"
    wb.save(out_file)
    print(f"Blank Excel workbook generated successfully: {out_file}")

if __name__ == "__main__":
    build_blank_assessment_excel()
