import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def build_lifelong_excel(is_blank=False):
    wb = openpyxl.Workbook()
    default_sheet = wb.active
    
    # Colors
    c_primary = "312E81"      # Deep Indigo
    c_primary_light = "E0E7FF"
    c_accent = "4338CA"       # Indigo
    c_accent_light = "EEF2FF"
    c_subtle = "F8FAFC"
    c_border = "CBD5E1"
    c_green_light = "DCFCE7"

    thin_border = Border(
        left=Side(style='thin', color=c_border),
        right=Side(style='thin', color=c_border),
        top=Side(style='thin', color=c_border),
        bottom=Side(style='thin', color=c_border)
    )
    thick_bottom = Border(
        left=Side(style='thin', color=c_border),
        right=Side(style='thin', color=c_border),
        top=Side(style='thin', color=c_border),
        bottom=Side(style='medium', color=c_primary)
    )

    # -------------------------------------------------------------
    # 1. Guide Sheet
    # -------------------------------------------------------------
    ws_guide = wb.create_sheet(title="評分標準與說明")
    ws_guide.views.sheetView[0].showGridLines = True
    
    suffix = " (空白範本)" if is_blank else " (含範例分析)"
    ws_guide["A1"] = f"瑞翔社區日間作業所 - 116年度終身學習領域四季主題學習檢核表{suffix}"
    ws_guide["A1"].font = Font(name="微軟正黑體", size=15, bold=True, color="FFFFFF")
    ws_guide["A1"].fill = PatternFill(start_color=c_primary, end_color=c_primary, fill_type="solid")
    ws_guide["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_guide.merge_cells("A1:F2")

    headers = ["分數", "提示等級代碼", "評分等級判定標準", "自立能力階段", "代表色彩", "個別化支持策略"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws_guide.cell(row=4, column=col_idx, value=h)
        cell.font = Font(name="微軟正黑體", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=c_accent, end_color=c_accent, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    standards = [
        (5, "獨立完成 (＋)", "不需要任何提示，能主動且正確無誤地完成完整工作分析步驟。", "🟢 獨立完成 (4.5~5.0分)", "綠色", "達到完全自立水準，可擔任同儕示範小幫手"),
        (4, "口語提醒 (Ｖ)", "支持者給予口頭提醒（如：「線材要怎麼收？」）即可正確完成。", "🔵 口語提醒 (3.5~4.4分)", "藍色", "達到基本自立達標標準，需持續維持習慣"),
        (3, "肢體協助 (Ｐ)", "需要支持者給予部分肢體引導或輕碰示範（如手把手帶領）才能完成。", "🟡 肢體協助 (2.5~3.4分)", "黃色", "需日常持續操作練習，拆解難點步驟"),
        (2, "部分協助 (Ｐ)", "需支持者協助操作超過一半或關鍵步驟才能完成。", "🟡 部分協助 (2.0~2.4分)", "橙色", "需個別化結構化圖卡輔助"),
        (1, "尚未學會 (－)", "大部分步驟由支持者代為操作，使用者僅能參與極少部分或大量協助。", "🟠 尚未學會 (1.0~1.9分)", "紅橙色", "需支持者全程密集陪伴與動機引導"),
        (0, "無法配合 (－)", "即使提供大量協助仍無法完成或拒絕配合操作。", "🔴 引導探索 (0.0~0.9分)", "紅色", "需建立操作意願與環境適應")
    ]

    for r_idx, row_data in enumerate(standards, 5):
        for c_idx, val in enumerate(row_data, 1):
            cell = ws_guide.cell(row=r_idx, column=c_idx, value=val)
            cell.font = Font(name="微軟正黑體", size=10)
            cell.alignment = Alignment(horizontal="center" if c_idx in [1, 2, 4, 5] else "left", vertical="center", wrap_text=True)
            cell.border = thin_border
            if r_idx % 2 == 0:
                cell.fill = PatternFill(start_color=c_subtle, end_color=c_subtle, fill_type="solid")

    ws_guide.column_dimensions['A'].width = 10
    ws_guide.column_dimensions['B'].width = 16
    ws_guide.column_dimensions['C'].width = 38
    ws_guide.column_dimensions['D'].width = 24
    ws_guide.column_dimensions['E'].width = 12
    ws_guide.column_dimensions['F'].width = 35

    # -------------------------------------------------------------
    # 2. Quarters Q1~Q4
    # -------------------------------------------------------------
    quarters_data = [
        {
            "id": "Q1",
            "title": "第一季 (1~3月)：生活常規、作息適應與自我倡議",
            "items": [
                "1. 能辨識小作所每日作息、週課表與停班停課資訊（如週六日休息、看新聞確認颱風假）",
                "2. 理解作業活動與獎勵金制度之關聯（知道認真代工/手作可獲取獎勵金用於社區購物）",
                "3. 能主動表達個人身心需求、身體不適或善用意見箱與支持者/社工溝通傾訴"
            ]
        },
        {
            "id": "Q2",
            "title": "第二季 (4~6月)：居家清潔維護與家電用電安全",
            "items": [
                "1. 掌握正確拔插頭技巧（握住插頭本體拔除，嚴禁直接拉扯電線）與線材束帶收納",
                "2. 能辨識用電危險（插座過載堆疊、插頭未插到底、電線脫皮破損立即通報拔除）",
                "3. 辨識廚房高溫熱源（電陶爐發紅面板會燙傷不可碰觸）與高耗能電器插座安全",
                "4. 熟練「麻花式」低處擰抹布技巧，並能使用掃帚將垃圾確實掃入地面目標框內"
            ]
        },
        {
            "id": "Q3",
            "title": "第三季 (7~9月)：全方位防災應變（地震、火災、水災）與用路安全",
            "items": [
                "1. 熟練地震避難口訣「趴下、掩護、穩住（躲、遮、抓）」，確實躲入桌下抓穩桌腳",
                "2. 能於非桌區（如作業區玻璃旁）迅速遠離玻璃、尋找柱子避難，睡覺時以枕頭護頭",
                "3. 聽聞火警警報迅速關閉教室門、用毛巾塞門縫防煙，並確實往陽台/室外避難",
                "4. 能正確操作滅火器四步驟「拉（插梢）、瞄（火源根部）、壓（握把）、掃（左右掃射）」",
                "5. 能辨識用路危險（路邊車輛遮蔽視線、無號誌路口、保持行走動線通暢）"
            ]
        },
        {
            "id": "Q4",
            "title": "第四季 (10~12月)：食品安全守則、期限判讀與生活消費管理",
            "items": [
                "1. 能於包裝上找出「EXP」、「有效日期」或「保存期限」標籤位置",
                "2. 能對照月曆比對當日日期，正確判斷食品或用品是否過期，過期主動丟棄不食用",
                "3. 掌握油鍋起火應變三步驟（1關瓦斯、2蓋鍋蓋、3靜置冷卻，嚴禁潑水）",
                "4. 練習日常生活購物金錢清點（如辨識百元/千元鈔、核對找零）並養成收納習慣"
            ]
        }
    ]

    learners_names = ["宇彤", "育萱", "高齊", "芷嫻", "志豪", "雅婷"] if not is_blank else [f"學員 {i+1}" for i in range(6)]

    sample_scores = {
        "Q1": {
            "宇彤": ([3,3,3], [4,4,4], [5,4,5]),
            "育萱": ([4,4,4], [4,5,5], [5,5,5]),
            "高齊": ([2,2,2], [3,3,4], [4,4,4]),
            "芷嫻": ([3,3,4], [4,4,4], [5,4,5]),
            "志豪": ([3,2,3], [4,3,4], [4,4,4]),
            "雅婷": ([3,3,3], [4,4,4], [5,5,4])
        },
        "Q2": {
            "宇彤": ([3,3,2,3], [4,4,3,4], [5,4,4,5]),
            "育萱": ([4,4,3,4], [5,5,4,5], [5,5,5,5]),
            "高齊": ([2,2,2,3], [3,3,3,4], [4,4,4,4]),
            "芷嫻": ([3,3,3,4], [4,4,4,5], [5,5,4,5]),
            "志豪": ([2,3,2,3], [3,4,3,4], [4,4,4,4]),
            "雅婷": ([3,3,3,3], [4,4,4,4], [5,5,4,5])
        },
        "Q3": {
            "宇彤": ([3,3,3,3,3], [4,4,4,4,4], [5,4,5,4,4]),
            "育萱": ([4,4,4,4,4], [5,5,5,5,4], [5,5,5,5,5]),
            "高齊": ([2,2,2,2,3], [3,3,3,4,3], [4,4,4,4,4]),
            "芷嫻": ([3,3,4,3,4], [4,4,5,4,4], [5,5,5,4,5]),
            "志豪": ([2,3,2,3,3], [3,4,3,4,4], [4,4,4,4,4]),
            "雅婷": ([3,3,3,4,3], [4,4,4,4,4], [5,5,4,5,5])
        },
        "Q4": {
            "宇彤": ([3,3,3,3], [4,4,4,4], [5,5,4,4]),
            "育萱": ([4,4,4,4], [5,5,5,5], [5,5,5,5]),
            "高齊": ([2,2,3,2], [3,4,3,3], [4,4,4,4]),
            "芷嫻": ([3,4,3,4], [4,4,4,5], [5,5,4,5]),
            "志豪": ([2,3,3,2], [3,4,4,3], [4,4,4,4]),
            "雅婷": ([3,3,4,3], [4,4,4,4], [5,5,4,5])
        }
    }

    for q_info in quarters_data:
        ws_q = wb.create_sheet(title=f"{q_info['id']}評量打分")
        ws_q.views.sheetView[0].showGridLines = True
        n_items = len(q_info['items'])
        max_score = n_items * 5

        # Title
        last_col_let = get_column_letter(2 + n_items + 7)
        ws_q["A1"] = f"瑞翔日間作業所 116年度終身學習 - {q_info['title']}"
        ws_q["A1"].font = Font(name="微軟正黑體", size=14, bold=True, color="FFFFFF")
        ws_q["A1"].fill = PatternFill(start_color=c_primary, end_color=c_primary, fill_type="solid")
        ws_q["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws_q.merge_cells(f"A1:{last_col_let}2")

        # Instructions banner
        ws_q["A3"] = "【提示等級說明】5:獨立完成(＋) ｜ 4:口語提醒(Ｖ) ｜ 3:肢體協助(Ｐ) ｜ 2:部分協助 ｜ 1:尚未學會(－) ｜ 0:無法配合"
        ws_q["A3"].font = Font(name="微軟正黑體", size=10, bold=True, color="312E81")
        ws_q["A3"].fill = PatternFill(start_color=c_primary_light, end_color=c_primary_light, fill_type="solid")
        ws_q["A3"].alignment = Alignment(horizontal="left", vertical="center")
        ws_q.merge_cells(f"A3:{last_col_let}3")

        # Table Header
        ws_q["A4"] = "學員姓名"
        ws_q["A4"].font = Font(name="微軟正黑體", size=11, bold=True, color="FFFFFF")
        ws_q["A4"].fill = PatternFill(start_color=c_accent, end_color=c_accent, fill_type="solid")
        ws_q["A4"].alignment = Alignment(horizontal="center", vertical="center")
        ws_q.merge_cells("A4:A5")

        ws_q["B4"] = "評量階段"
        ws_q["B4"].font = Font(name="微軟正黑體", size=11, bold=True, color="FFFFFF")
        ws_q["B4"].fill = PatternFill(start_color=c_accent, end_color=c_accent, fill_type="solid")
        ws_q["B4"].alignment = Alignment(horizontal="center", vertical="center")
        ws_q.merge_cells("B4:B5")

        for i in range(1, n_items + 1):
            col_letter = get_column_letter(2 + i)
            cell = ws_q[f"{col_letter}4"]
            cell.value = f"指標 {i}"
            cell.font = Font(name="微軟正黑體", size=10, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=c_accent, end_color=c_accent, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

            sub_cell = ws_q[f"{col_letter}5"]
            sub_cell.value = f"(0~5分)"
            sub_cell.font = Font(name="微軟正黑體", size=8, color="FFFFFF")
            sub_cell.fill = PatternFill(start_color=c_accent, end_color=c_accent, fill_type="solid")
            sub_cell.alignment = Alignment(horizontal="center", vertical="center")

        calc_start_col = 2 + n_items + 1
        calc_headers = [
            ("總得分", f"(滿分{max_score})"),
            ("平均分", "(0~5分)"),
            ("學習自立階段", "(等級標記)"),
            ("達成率", "(%)"),
            ("進步分數", "(後-前)"),
            ("進步率", "(%)"),
            ("成長百分點", "(%pt)")
        ]
        for offset, (h_main, h_sub) in enumerate(calc_headers):
            col_let = get_column_letter(calc_start_col + offset)
            c1 = ws_q[f"{col_let}4"]
            c1.value = h_main
            c1.font = Font(name="微軟正黑體", size=10, bold=True, color="FFFFFF")
            c1.fill = PatternFill(start_color="1E1B4B", end_color="1E1B4B", fill_type="solid")
            c1.alignment = Alignment(horizontal="center", vertical="center")

            c2 = ws_q[f"{col_let}5"]
            c2.value = h_sub
            c2.font = Font(name="微軟正黑體", size=8, color="FFFFFF")
            c2.fill = PatternFill(start_color="1E1B4B", end_color="1E1B4B", fill_type="solid")
            c2.alignment = Alignment(horizontal="center", vertical="center")

        start_row = 6
        last_score_col_let = get_column_letter(2 + n_items)
        col_sum_let = get_column_letter(calc_start_col)
        col_avg_let = get_column_letter(calc_start_col + 1)
        col_stg_let = get_column_letter(calc_start_col + 2)
        col_ach_let = get_column_letter(calc_start_col + 3)
        col_gsc_let = get_column_letter(calc_start_col + 4)
        col_grt_let = get_column_letter(calc_start_col + 5)
        col_gpt_let = get_column_letter(calc_start_col + 6)

        for l_idx, learner_name in enumerate(learners_names):
            r_pre = start_row + l_idx * 3
            r_mid = r_pre + 1
            r_post = r_pre + 2

            ws_q[f"A{r_pre}"] = learner_name
            ws_q[f"A{r_pre}"].font = Font(name="微軟正黑體", size=11, bold=True)
            ws_q[f"A{r_pre}"].alignment = Alignment(horizontal="center", vertical="center")
            ws_q.merge_cells(f"A{r_pre}:A{r_post}")

            ws_q[f"B{r_pre}"] = "1. 前測 (第1月)"
            ws_q[f"B{r_mid}"] = "2. 課堂 (第2月)"
            ws_q[f"B{r_post}"] = "3. 後測 (第3月)"
            for r, stage_fill in [(r_pre, "F8FAFC"), (r_mid, "EEF2FF"), (r_post, "E0E7FF")]:
                c = ws_q[f"B{r}"]
                c.font = Font(name="微軟正黑體", size=9, bold=(r == r_post))
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.fill = PatternFill(start_color=stage_fill, end_color=stage_fill, fill_type="solid")

            # Scores
            sample_tuple = sample_scores[q_info["id"]].get(learner_name, ([0]*n_items, [0]*n_items, [0]*n_items))
            for i in range(n_items):
                col_let = get_column_letter(3 + i)
                for r, s_idx in [(r_pre, 0), (r_mid, 1), (r_post, 2)]:
                    cell = ws_q[f"{col_let}{r}"]
                    cell.value = "" if is_blank else sample_tuple[s_idx][i]
                    cell.font = Font(name="微軟正黑體", size=10)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border

            # Formulas
            for r in [r_pre, r_mid, r_post]:
                ws_q[f"{col_sum_let}{r}"] = f'=IF(COUNT(C{r}:{last_score_col_let}{r})>0, SUM(C{r}:{last_score_col_let}{r}), "")'
                ws_q[f"{col_avg_let}{r}"] = f'=IF(COUNT(C{r}:{last_score_col_let}{r})>0, AVERAGE(C{r}:{last_score_col_let}{r}), "")'
                ws_q[f"{col_stg_let}{r}"] = f'=IF({col_avg_let}{r}="","",IF({col_avg_let}{r}>=4.5,"🟢 獨立完成",IF({col_avg_let}{r}>=3.5,"🔵 口語提醒",IF({col_avg_let}{r}>=2.0,"🟡 肢體/部分協助","🟠 尚未學會"))))'
                ws_q[f"{col_ach_let}{r}"] = f'=IF(ISNUMBER({col_sum_let}{r}), {col_sum_let}{r}/{max_score}, "")'

                ws_q[f"{col_sum_let}{r}"].font = Font(name="微軟正黑體", size=10, bold=True)
                ws_q[f"{col_avg_let}{r}"].font = Font(name="微軟正黑體", size=10)
                ws_q[f"{col_stg_let}{r}"].font = Font(name="微軟正黑體", size=9, bold=True)
                ws_q[f"{col_ach_let}{r}"].font = Font(name="微軟正黑體", size=10)
                ws_q[f"{col_avg_let}{r}"].number_format = '0.0'
                ws_q[f"{col_ach_let}{r}"].number_format = '0.0%'

                for col in [col_sum_let, col_avg_let, col_stg_let, col_ach_let]:
                    ws_q[f"{col}{r}"].alignment = Alignment(horizontal="center", vertical="center")
                    ws_q[f"{col}{r}"].border = thin_border

            # Growth formulas
            ws_q[f"{col_gsc_let}{r_pre}"] = f'=IF(AND(ISNUMBER({col_sum_let}{r_pre}), ISNUMBER({col_sum_let}{r_post})), {col_sum_let}{r_post}-{col_sum_let}{r_pre}, "")'
            ws_q[f"{col_grt_let}{r_pre}"] = f'=IF(AND(ISNUMBER({col_sum_let}{r_pre}), ISNUMBER({col_sum_let}{r_post})), IF({col_sum_let}{r_pre}>0, ({col_sum_let}{r_post}-{col_sum_let}{r_pre})/{col_sum_let}{r_pre}, ({col_sum_let}{r_post}-{col_sum_let}{r_pre})/{max_score}), "")'
            ws_q[f"{col_gpt_let}{r_pre}"] = f'=IF(AND(ISNUMBER({col_ach_let}{r_pre}), ISNUMBER({col_ach_let}{r_post})), {col_ach_let}{r_post}-{col_ach_let}{r_pre}, "")'

            ws_q.merge_cells(f"{col_gsc_let}{r_pre}:{col_gsc_let}{r_post}")
            ws_q.merge_cells(f"{col_grt_let}{r_pre}:{col_grt_let}{r_post}")
            ws_q.merge_cells(f"{col_gpt_let}{r_pre}:{col_gpt_let}{r_post}")

            for col in [col_gsc_let, col_grt_let, col_gpt_let]:
                cell = ws_q[f"{col}{r_pre}"]
                cell.font = Font(name="微軟正黑體", size=10, bold=True, color="166534")
                cell.fill = PatternFill(start_color=c_green_light, end_color=c_green_light, fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

            ws_q[f"{col_gsc_let}{r_pre}"].number_format = '+0;-0;0'
            ws_q[f"{col_grt_let}{r_pre}"].number_format = '+0.0%;-0.0%;0.0%'
            ws_q[f"{col_gpt_let}{r_pre}"].number_format = '+0.0%;-0.0%;0.0%'

            for col_idx in range(1, calc_start_col + 7):
                cell_name = get_column_letter(col_idx)
                for r in [r_pre, r_mid, r_post]:
                    ws_q[f"{cell_name}{r}"].border = thin_border

        # Averages row
        r_avg = start_row + len(learners_names) * 3
        ws_q[f"A{r_avg}"] = "全班平均常模"
        ws_q[f"A{r_avg}"].font = Font(name="微軟正黑體", size=10, bold=True, color="FFFFFF")
        ws_q[f"A{r_avg}"].fill = PatternFill(start_color=c_primary, end_color=c_primary, fill_type="solid")
        ws_q[f"A{r_avg}"].alignment = Alignment(horizontal="center", vertical="center")
        ws_q.merge_cells(f"A{r_avg}:B{r_avg}")

        for i in range(n_items):
            col_let = get_column_letter(3 + i)
            cell = ws_q[f"{col_let}{r_avg}"]
            cell.value = f'=IFERROR(AVERAGE({col_let}6:{col_let}{r_avg-1}), "-")'
            cell.font = Font(name="微軟正黑體", size=10, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(start_color=c_primary_light, end_color=c_primary_light, fill_type="solid")
            cell.border = thick_bottom

        ws_q[f"{col_sum_let}{r_avg}"] = f'=IFERROR(AVERAGE({col_sum_let}6:{col_sum_let}{r_avg-1}), "-")'
        ws_q[f"{col_avg_let}{r_avg}"] = f'=IFERROR(AVERAGE({col_avg_let}6:{col_avg_let}{r_avg-1}), "-")'
        ws_q[f"{col_stg_let}{r_avg}"] = "-"
        ws_q[f"{col_ach_let}{r_avg}"] = f'=IFERROR(AVERAGE({col_ach_let}6:{col_ach_let}{r_avg-1}), "-")'
        ws_q[f"{col_gsc_let}{r_avg}"] = f'=IFERROR(AVERAGE({col_gsc_let}6:{col_gsc_let}{r_avg-1}), "-")'
        ws_q[f"{col_grt_let}{r_avg}"] = f'=IFERROR(AVERAGE({col_grt_let}6:{col_grt_let}{r_avg-1}), "-")'
        ws_q[f"{col_gpt_let}{r_avg}"] = f'=IFERROR(AVERAGE({col_gpt_let}6:{col_gpt_let}{r_avg-1}), "-")'

        for col in [col_sum_let, col_avg_let, col_stg_let, col_ach_let, col_gsc_let, col_grt_let, col_gpt_let]:
            cell = ws_q[f"{col}{r_avg}"]
            cell.font = Font(name="微軟正黑體", size=10, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(start_color=c_primary_light, end_color=c_primary_light, fill_type="solid")
            cell.border = thick_bottom

        # Indicators Reference
        r_ref = r_avg + 3
        ws_q.cell(row=r_ref, column=1, value=f"【{q_info['id']} 檢核指標工作分析項目清單】").font = Font(name="微軟正黑體", size=11, bold=True, color=c_primary)
        for i, item_text in enumerate(q_info['items'], 1):
            r_item = r_ref + i
            ws_q.cell(row=r_item, column=1, value=f"指標 {i}").alignment = Alignment(horizontal="center")
            ws_q.cell(row=r_item, column=2, value=item_text)
            ws_q.merge_cells(start_row=r_item, start_column=2, end_row=r_item, end_column=calc_start_col + 6)
            for col in range(1, calc_start_col + 7):
                ws_q.cell(row=r_item, column=col).font = Font(name="微軟正黑體", size=9)
                ws_q.cell(row=r_item, column=col).border = thin_border

        # Widths
        ws_q.column_dimensions['A'].width = 14
        ws_q.column_dimensions['B'].width = 16
        for i in range(1, n_items + 1):
            ws_q.column_dimensions[get_column_letter(2 + i)].width = 9
        ws_q.column_dimensions[col_sum_let].width = 10
        ws_q.column_dimensions[col_avg_let].width = 9
        ws_q.column_dimensions[col_stg_let].width = 18
        ws_q.column_dimensions[col_ach_let].width = 10
        ws_q.column_dimensions[col_gsc_let].width = 11
        ws_q.column_dimensions[col_grt_let].width = 11
        ws_q.column_dimensions[col_gpt_let].width = 13

    if default_sheet in wb.worksheets:
        wb.remove(default_sheet)

    out_file = "116年度終身學習四季主題檢核表_空白範本檔.xlsx" if is_blank else "116年度終身學習四季主題檢核表.xlsx"
    wb.save(out_file)
    print(f"Lifelong Excel saved: {out_file}")

if __name__ == "__main__":
    build_lifelong_excel(is_blank=False)
    build_lifelong_excel(is_blank=True)
